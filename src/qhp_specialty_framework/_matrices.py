"""Compatibility matrices (loaded from Excel or hardcoded)."""

from __future__ import annotations

from typing import Optional

import openpyxl


class CompatibilityMatrices:
    """Loads and provides access to the specialty compatibility matrices."""

    def __init__(self, excel_path: Optional[str] = None):
        if excel_path:
            self._load_from_excel(excel_path)
        else:
            self._load_hardcoded()

    def _load_from_excel(self, path: str) -> None:
        """Load matrices from the CMS Excel file."""
        wb = openpyxl.load_workbook(path)

        # Phys Specialty Compat Matrix: rows 4-12, cols C-K
        ws = wb["Phys Specialty Compat Matrix"]
        self._specialty_compat: dict[tuple[str, str], bool] = {}

        col_headers = []
        for col in range(3, 12):
            val = ws.cell(row=3, column=col).value
            if val:
                code = str(val).split()[0]
                col_headers.append(code)

        for row in range(4, 13):
            row_code_cell = ws.cell(row=row, column=2).value
            if not row_code_cell:
                continue
            row_code = str(row_code_cell).split()[0]
            for i, col_code in enumerate(col_headers):
                val = ws.cell(row=row, column=3 + i).value
                if val in ("Y", "y", "Yes", "yes"):
                    self._specialty_compat[(row_code, col_code)] = True
                else:
                    self._specialty_compat[(row_code, col_code)] = False

        # Phys Subspecialty Compat Matrix: rows 4-14, cols C-K
        ws = wb["Phys Subspecialty Compat Matrix"]
        self._subspecialty_compat: dict[tuple[str, str], bool] = {}

        col_headers = []
        for col in range(3, 12):
            val = ws.cell(row=3, column=col).value
            if val:
                code = str(val).split()[0]
                col_headers.append(code)

        for row in range(4, 15):
            row_code_cell = ws.cell(row=row, column=2).value
            if not row_code_cell:
                continue
            row_code = str(row_code_cell).split()[0]
            for i, col_code in enumerate(col_headers):
                val = ws.cell(row=row, column=3 + i).value
                if val in ("Y", "y", "Yes", "yes"):
                    self._subspecialty_compat[(row_code, col_code)] = True
                else:
                    self._subspecialty_compat[(row_code, col_code)] = False

    def _load_hardcoded(self) -> None:
        """Hardcoded matrices from the Excel file content."""
        specialty_compat_data = {
            # 002 Family Medicine
            ("002", "002"): True, ("002", "003"): True, ("002", "101"): False,
            ("002", "007"): True, ("002", "011"): False, ("002", "037"): True,
            ("002", "022"): False, ("002", "026"): False, ("002", "029"): True,
            # 003 Internal Medicine
            ("003", "002"): True, ("003", "003"): True, ("003", "101"): True,
            ("003", "007"): True, ("003", "011"): True, ("003", "037"): True,
            ("003", "022"): False, ("003", "026"): False, ("003", "029"): True,
            # 101 Primary Care - Pediatric
            ("101", "002"): False, ("101", "003"): True, ("101", "101"): True,
            ("101", "007"): True, ("101", "011"): True, ("101", "037"): True,
            ("101", "022"): False, ("101", "026"): True, ("101", "029"): True,
            # 007 Allergy and Immunology
            ("007", "002"): True, ("007", "003"): True, ("007", "101"): True,
            ("007", "007"): True, ("007", "011"): False, ("007", "037"): False,
            ("007", "022"): False, ("007", "026"): False, ("007", "029"): False,
            # 011 Dermatology
            ("011", "002"): False, ("011", "003"): True, ("011", "101"): True,
            ("011", "007"): False, ("011", "011"): True, ("011", "037"): False,
            ("011", "022"): False, ("011", "026"): False, ("011", "029"): False,
            # 037 Emergency Medicine
            ("037", "002"): True, ("037", "003"): True, ("037", "101"): True,
            ("037", "007"): False, ("037", "011"): False, ("037", "037"): True,
            ("037", "022"): False, ("037", "026"): False, ("037", "029"): False,
            # 022 Oncology - Radiation
            ("022", "002"): False, ("022", "003"): False, ("022", "101"): False,
            ("022", "007"): False, ("022", "011"): False, ("022", "037"): False,
            ("022", "022"): True, ("022", "026"): False, ("022", "029"): False,
            # 026 Physical Medicine & Rehabilitation
            ("026", "002"): False, ("026", "003"): False, ("026", "101"): True,
            ("026", "007"): False, ("026", "011"): False, ("026", "037"): False,
            ("026", "022"): False, ("026", "026"): True, ("026", "029"): False,
            # 029 Psychiatry
            ("029", "002"): True, ("029", "003"): True, ("029", "101"): True,
            ("029", "007"): False, ("029", "011"): False, ("029", "037"): False,
            ("029", "022"): False, ("029", "026"): False, ("029", "029"): True,
        }
        self._specialty_compat = specialty_compat_data

        subspecialty_compat_data = {
            # 008 Cardiology
            ("008", "002"): False, ("008", "003"): True, ("008", "101"): True,
            ("008", "007"): False, ("008", "011"): False, ("008", "037"): False,
            ("008", "022"): False, ("008", "026"): False, ("008", "029"): False,
            # 012 Endocrinology
            ("012", "002"): False, ("012", "003"): True, ("012", "101"): True,
            ("012", "007"): False, ("012", "011"): False, ("012", "037"): False,
            ("012", "022"): False, ("012", "026"): False, ("012", "029"): False,
            # 014 Gastroenterology
            ("014", "002"): False, ("014", "003"): True, ("014", "101"): True,
            ("014", "007"): False, ("014", "011"): False, ("014", "037"): False,
            ("014", "022"): False, ("014", "026"): False, ("014", "029"): False,
            # 004 Geriatrics
            ("004", "002"): True, ("004", "003"): True, ("004", "101"): False,
            ("004", "007"): False, ("004", "011"): False, ("004", "037"): False,
            ("004", "022"): False, ("004", "026"): False, ("004", "029"): True,
            # 017 Infectious Diseases
            ("017", "002"): False, ("017", "003"): True, ("017", "101"): True,
            ("017", "007"): False, ("017", "011"): False, ("017", "037"): False,
            ("017", "022"): False, ("017", "026"): False, ("017", "029"): False,
            # 018 Nephrology
            ("018", "002"): False, ("018", "003"): True, ("018", "101"): True,
            ("018", "007"): False, ("018", "011"): False, ("018", "037"): False,
            ("018", "022"): False, ("018", "026"): False, ("018", "029"): False,
            # 019 Neurology
            ("019", "002"): False, ("019", "003"): True, ("019", "101"): True,
            ("019", "007"): False, ("019", "011"): False, ("019", "037"): False,
            ("019", "022"): False, ("019", "026"): False, ("019", "029"): True,
            # 030 Pulmonology
            ("030", "002"): False, ("030", "003"): True, ("030", "101"): True,
            ("030", "007"): False, ("030", "011"): False, ("030", "037"): False,
            ("030", "022"): False, ("030", "026"): False, ("030", "029"): False,
            # 031 Rheumatology
            ("031", "002"): False, ("031", "003"): True, ("031", "101"): True,
            ("031", "007"): False, ("031", "011"): False, ("031", "037"): False,
            ("031", "022"): False, ("031", "026"): False, ("031", "029"): False,
            # 800 Addiction Medicine Physician
            ("800", "002"): True, ("800", "003"): True, ("800", "101"): True,
            ("800", "007"): True, ("800", "011"): True, ("800", "037"): True,
            ("800", "022"): True, ("800", "026"): True, ("800", "029"): True,
            # 021 Oncology - Medical & Surgical
            ("021", "002"): False, ("021", "003"): True, ("021", "101"): True,
            ("021", "007"): False, ("021", "011"): False, ("021", "037"): False,
            ("021", "022"): False, ("021", "026"): False, ("021", "029"): False,
        }
        self._subspecialty_compat = subspecialty_compat_data

    def are_specialties_compatible(self, code_a: str, code_b: str) -> bool:
        """Check if two physician specialties are compatible (Y/N from matrix)."""
        key = (code_a, code_b)
        if key in self._specialty_compat:
            return self._specialty_compat[key]
        return code_a == code_b

    def is_subspecialty_compatible_with_specialty(
        self, subspecialty_code: str, specialty_code: str
    ) -> bool:
        """Check if a subspecialty is compatible with a specialty."""
        key = (subspecialty_code, specialty_code)
        if key in self._subspecialty_compat:
            return self._subspecialty_compat[key]
        return False
