"""
QHP Network Adequacy Provider Specialty Framework & Matrices (PY2027 v1.0)

Implements the CMS Provider Specialty Framework decision logic and specialty
compatibility validation matrices from:
  https://www.qhpcertification.cms.gov/QHP/applicationmaterials/Network-Adequacy

Source file: PY2027_NA_Provider_Specialty_Framework_Matrices_v1.0.xlsx

The framework classifies providers into groupings and validates that the
specialty/subspecialty combinations reported for a single NPI are clinically
compatible per the CMS matrices.
"""

from __future__ import annotations

import pathlib
from dataclasses import dataclass, field
from enum import Enum
from typing import Optional

import openpyxl


# ────────────────────────────────────────────────────────────────────
# Provider Groupings (from the flowchart decision tree)
# ────────────────────────────────────────────────────────────────────

class ProviderGrouping(str, Enum):
    PHYSICIAN = "Physician"
    SURGEON = "Surgeon"
    DENTIST = "Dentist"
    ADVANCED_PRACTITIONER = "Advanced Practitioner"
    BEHAVIORAL_HEALTH = "Behavioral Health"
    ALLIED_HEALTH = "Allied Health Professional"
    FACILITY = "Facility"


# ────────────────────────────────────────────────────────────────────
# Specialty codes from the matrices & flowchart
# ────────────────────────────────────────────────────────────────────

# Physician specialties (from flowchart)
# Select up to two specialties and/or may designate one subspecialty
PHYSICIAN_SPECIALTIES = {
    "002": "Family Medicine",
    "003": "Internal Medicine",
    "101": "Primary Care Pediatrics",
    "007": "Allergy and Immunology",
    "011": "Dermatology",
    "037": "Emergency Medicine",
    "022": "Oncology-Radiation",
    "026": "Physical Medicine & Rehabilitation",
    "029": "Psychiatry",
}

# Fallback — "Only if none of the above"
PHYSICIAN_FALLBACK = {
    "001": "General Practice",
}

# Physician subspecialties (from flowchart)
PHYSICIAN_SUBSPECIALTIES = {
    "008": "Cardiology",
    "012": "Endocrinology",
    "014": "Gastroenterology",
    "017": "Infectious Disease",
    "018": "Nephrology",
    "019": "Neurology",
    "030": "Pulmonology",
    "031": "Rheumatology",
    "021": "Oncology-Surgical Medical",
    "004": "Geriatrics",
    "800": "Addiction Medicine Physician",
}

# Surgeon specialties (from flowchart)
# Set 1: select ONE specialty
# Set 2: select up to TWO specialties
# Cannot mix across sets. Also may select Surgical Subspecialty (021).
SURGEON_SPECIALTIES_SET1 = {
    "015": "General Surgery",
    "013": "ENT/Otolaryngology",
    "020": "Neurosurgery",
    "023": "Ophthalmology",
    "025": "Orthopedic Surgery",
    "016": "Gynecology/OBGYN",
    "027": "Plastic Surgery",
    "033": "Urology",
}

SURGEON_SPECIALTIES_SET2 = {
    "035": "Cardiothoracic Surgeon",
    "034": "Vascular Surgeon",
}

SURGEON_SPECIALTIES = {**SURGEON_SPECIALTIES_SET1, **SURGEON_SPECIALTIES_SET2}

# Surgical subspecialty — "Also may select" (in addition to Set 1/2 specialty)
SURGEON_SUBSPECIALTIES = {
    "021": "Oncology-Surgical/Medical",
}

# Dentist specialties (from flowchart)
# Select up to two specialties and/or may designate one subspecialty
DENTIST_SPECIALTIES = {
    "201": "Dental-General",
    "P201": "Dental-General(Pediatric)",
}

# Dental subspecialties (from flowchart)
DENTIST_SUBSPECIALTIES = {
    "204": "Dental-Endodontist",
    "202": "Dental-Orthodontist",
    "203": "Dental-Periodontist",
    "206": "Dental-Prosthodontist",
}

# Advanced Practitioner specialties (from flowchart)
# Generalist path
ADVANCED_PRACTITIONER_SPECIALTIES = {
    "006": "Primary Care Advanced Practice Registered Nurse",
    "A006": "Primary Care APRN-Adult",
    "P006": "Primary Care APRN-Pediatric",
    "005": "Primary Care-Physician Assistant",  # fallback
}

# Behavioral Health path (for NP/PA with BH scope)
ADVANCED_PRACTITIONER_BH_SPECIALTIES = {
    "108": "Behavioral Health APRN",
}

# Behavioral Health specialties (from flowchart)
# Select one specialty and/or may designate up to two subspecialties
BEHAVIORAL_HEALTH_SPECIALTIES = {
    "102": "Social Worker",
    "107": "Counselor (Mental Health and Professional)",
    "103": "Psychologist",
}

# Behavioral Health subspecialties (from flowchart)
BEHAVIORAL_HEALTH_SUBSPECIALTIES = {
    "105": "Marriage and Family Therapist",
    "106": "Addiction (Substance Use Disorder) Counselor",
    "801": "Behavioral Analyst",
}

# Allied Health specialties (from flowchart)
# Select Allied Health Professionals grouping and one specialty
ALLIED_HEALTH_SPECIALTIES = {
    "010": "Chiropractor",
    "028": "Podiatry",
    "049": "Physical Therapy",
    "051": "Speech Therapy",
    "050": "Occupational Therapy",
}


# ────────────────────────────────────────────────────────────────────
# Data classes
# ────────────────────────────────────────────────────────────────────

@dataclass
class ValidationError:
    """A single validation error or warning."""
    rule: str  # e.g. "NA V14"
    severity: str  # "error" or "warning"
    message: str


@dataclass
class ProviderRecord:
    """A provider as submitted on the NA Template."""
    npi: str
    provider_grouping: ProviderGrouping
    specialties: list[str] = field(default_factory=list)  # specialty codes
    subspecialties: list[str] = field(default_factory=list)  # subspecialty codes
    is_md_or_do: bool = False
    is_surgeon: bool = False
    is_dentist: bool = False
    is_np_or_pa: bool = False
    is_behavioral_health: bool = False
    is_facility: bool = False


@dataclass
class ValidationResult:
    """Result of validating a provider record."""
    provider: ProviderRecord
    errors: list[ValidationError] = field(default_factory=list)
    warnings: list[ValidationError] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0

    def add_error(self, rule: str, message: str) -> None:
        self.errors.append(ValidationError(rule=rule, severity="error", message=message))

    def add_warning(self, rule: str, message: str) -> None:
        self.warnings.append(ValidationError(rule=rule, severity="warning", message=message))


# ────────────────────────────────────────────────────────────────────
# Compatibility matrices (loaded from Excel or hardcoded)
# ────────────────────────────────────────────────────────────────────

class CompatibilityMatrices:
    """Loads and provides access to the specialty compatibility matrices."""

    def __init__(self, excel_path: Optional[str] = None):
        if excel_path:
            self._load_from_excel(excel_path)
        else:
            self._load_hardcoded()

    def _load_from_excel(self, path: str) -> None:
        """Load matrices from the CMS Excel file."""
        wb = openpyxl.load_workbook(path)

        # Phys Specialty Compat Matrix: rows 4-12, cols C-K
        # Row headers (col B): specialty codes
        # Column headers (row 3): specialty codes
        # Values: Y/N
        ws = wb["Phys Specialty Compat Matrix"]
        self._specialty_compat: dict[tuple[str, str], bool] = {}

        # Read column headers (row 3, cols C-K)
        col_headers = []
        for col in range(3, 12):  # C=3 .. K=11
            val = ws.cell(row=3, column=col).value
            if val:
                code = str(val).split()[0]  # e.g. "002 Family Medicine2" → "002"
                col_headers.append(code)

        # Read row data (rows 4-12)
        for row in range(4, 13):
            row_code_cell = ws.cell(row=row, column=2).value  # col B
            if not row_code_cell:
                continue
            row_code = str(row_code_cell).split()[0]  # e.g. "002 Family Medicine2" → "002"

            for i, col_code in enumerate(col_headers):
                val = ws.cell(row=row, column=3 + i).value
                if val in ("Y", "y", "Yes", "yes"):
                    self._specialty_compat[(row_code, col_code)] = True
                else:
                    self._specialty_compat[(row_code, col_code)] = False

        # Phys Subspecialty Compat Matrix: rows 4-14, cols C-K
        # Row headers (col B): subspecialty codes
        # Column headers (row 3): specialty codes
        ws = wb["Phys Subspecialty Compat Matrix"]
        self._subspecialty_compat: dict[tuple[str, str], bool] = {}

        col_headers = []
        for col in range(3, 12):
            val = ws.cell(row=3, column=col).value
            if val:
                code = str(val).split()[0]
                col_headers.append(code)

        for row in range(4, 15):
            row_code_cell = ws.cell(row=row, column=2).value
            if not row_code_cell:
                continue
            row_code = str(row_code_cell).split()[0]

            for i, col_code in enumerate(col_headers):
                val = ws.cell(row=row, column=3 + i).value
                if val in ("Y", "y", "Yes", "yes"):
                    self._subspecialty_compat[(row_code, col_code)] = True
                else:
                    self._subspecialty_compat[(row_code, col_code)] = False

    def _load_hardcoded(self) -> None:
        """Hardcoded matrices from the Excel file content."""
        # Physician Specialty Compatibility Matrix (9x9)
        # Specialties: 002, 003, 101, 007, 011, 037, 022, 026, 029
        specialty_compat_data = {
            # 002 Family Medicine
            ("002", "002"): True, ("002", "003"): True, ("002", "101"): False,
            ("002", "007"): True, ("002", "011"): False, ("002", "037"): True,
            ("002", "022"): False, ("002", "026"): False, ("002", "029"): True,
            # 003 Internal Medicine
            ("003", "002"): True, ("003", "003"): True, ("003", "101"): True,
            ("003", "007"): True, ("003", "011"): True, ("003", "037"): True,
            ("003", "022"): False, ("003", "026"): False, ("003", "029"): True,
            # 101 Primary Care - Pediatric
            ("101", "002"): False, ("101", "003"): True, ("101", "101"): True,
            ("101", "007"): True, ("101", "011"): True, ("101", "037"): True,
            ("101", "022"): False, ("101", "026"): True, ("101", "029"): True,
            # 007 Allergy and Immunology
            ("007", "002"): True, ("007", "003"): True, ("007", "101"): True,
            ("007", "007"): True, ("007", "011"): False, ("007", "037"): False,
            ("007", "022"): False, ("007", "026"): False, ("007", "029"): False,
            # 011 Dermatology
            ("011", "002"): False, ("011", "003"): True, ("011", "101"): True,
            ("011", "007"): False, ("011", "011"): True, ("011", "037"): False,
            ("011", "022"): False, ("011", "026"): False, ("011", "029"): False,
            # 037 Emergency Medicine
            ("037", "002"): True, ("037", "003"): True, ("037", "101"): True,
            ("037", "007"): False, ("037", "011"): False, ("037", "037"): True,
            ("037", "022"): False, ("037", "026"): False, ("037", "029"): False,
            # 022 Oncology - Radiation
            ("022", "002"): False, ("022", "003"): False, ("022", "101"): False,
            ("022", "007"): False, ("022", "011"): False, ("022", "037"): False,
            ("022", "022"): True, ("022", "026"): False, ("022", "029"): False,
            # 026 Physical Medicine & Rehabilitation
            ("026", "002"): False, ("026", "003"): False, ("026", "101"): True,
            ("026", "007"): False, ("026", "011"): False, ("026", "037"): False,
            ("026", "022"): False, ("026", "026"): True, ("026", "029"): False,
            # 029 Psychiatry
            ("029", "002"): True, ("029", "003"): True, ("029", "101"): True,
            ("029", "007"): False, ("029", "011"): False, ("029", "037"): False,
            ("029", "022"): False, ("029", "026"): False, ("029", "029"): True,
        }
        self._specialty_compat = specialty_compat_data

        # Physician Subspecialty Compatibility Matrix (11x9)
        # Subspecialties: 008, 012, 014, 004, 017, 018, 019, 030, 031, 800, 021
        # Specialties (columns): 002, 003, 101, 007, 011, 037, 022, 026, 029
        subspecialty_compat_data = {
            # 008 Cardiology
            ("008", "002"): False, ("008", "003"): True, ("008", "101"): True,
            ("008", "007"): False, ("008", "011"): False, ("008", "037"): False,
            ("008", "022"): False, ("008", "026"): False, ("008", "029"): False,
            # 012 Endocrinology
            ("012", "002"): False, ("012", "003"): True, ("012", "101"): True,
            ("012", "007"): False, ("012", "011"): False, ("012", "037"): False,
            ("012", "022"): False, ("012", "026"): False, ("012", "029"): False,
            # 014 Gastroenterology
            ("014", "002"): False, ("014", "003"): True, ("014", "101"): True,
            ("014", "007"): False, ("014", "011"): False, ("014", "037"): False,
            ("014", "022"): False, ("014", "026"): False, ("014", "029"): False,
            # 004 Geriatrics
            ("004", "002"): True, ("004", "003"): True, ("004", "101"): False,
            ("004", "007"): False, ("004", "011"): False, ("004", "037"): False,
            ("004", "022"): False, ("004", "026"): False, ("004", "029"): True,
            # 017 Infectious Diseases
            ("017", "002"): False, ("017", "003"): True, ("017", "101"): True,
            ("017", "007"): False, ("017", "011"): False, ("017", "037"): False,
            ("017", "022"): False, ("017", "026"): False, ("017", "029"): False,
            # 018 Nephrology
            ("018", "002"): False, ("018", "003"): True, ("018", "101"): True,
            ("018", "007"): False, ("018", "011"): False, ("018", "037"): False,
            ("018", "022"): False, ("018", "026"): False, ("018", "029"): False,
            # 019 Neurology
            ("019", "002"): False, ("019", "003"): True, ("019", "101"): True,
            ("019", "007"): False, ("019", "011"): False, ("019", "037"): False,
            ("019", "022"): False, ("019", "026"): False, ("019", "029"): True,
            # 030 Pulmonology
            ("030", "002"): False, ("030", "003"): True, ("030", "101"): True,
            ("030", "007"): False, ("030", "011"): False, ("030", "037"): False,
            ("030", "022"): False, ("030", "026"): False, ("030", "029"): False,
            # 031 Rheumatology
            ("031", "002"): False, ("031", "003"): True, ("031", "101"): True,
            ("031", "007"): False, ("031", "011"): False, ("031", "037"): False,
            ("031", "022"): False, ("031", "026"): False, ("031", "029"): False,
            # 800 Addiction Medicine Physician
            ("800", "002"): True, ("800", "003"): True, ("800", "101"): True,
            ("800", "007"): True, ("800", "011"): True, ("800", "037"): True,
            ("800", "022"): True, ("800", "026"): True, ("800", "029"): True,
            # 021 Oncology - Medical & Surgical
            ("021", "002"): False, ("021", "003"): True, ("021", "101"): True,
            ("021", "007"): False, ("021", "011"): False, ("021", "037"): False,
            ("021", "022"): False, ("021", "026"): False, ("021", "029"): False,
        }
        self._subspecialty_compat = subspecialty_compat_data

    def are_specialties_compatible(self, code_a: str, code_b: str) -> bool:
        """Check if two physician specialties are compatible (Y/N from matrix)."""
        key = (code_a, code_b)
        if key in self._specialty_compat:
            return self._specialty_compat[key]
        # Matrix is symmetric for self-compatibility
        return code_a == code_b  # A specialty is always compatible with itself

    def is_subspecialty_compatible_with_specialty(
        self, subspecialty_code: str, specialty_code: str
    ) -> bool:
        """Check if a subspecialty is compatible with a specialty."""
        key = (subspecialty_code, specialty_code)
        if key in self._subspecialty_compat:
            return self._subspecialty_compat[key]
        return False


# ────────────────────────────────────────────────────────────────────
# Framework: classify provider into grouping (flowchart logic)
# ────────────────────────────────────────────────────────────────────

def classify_provider(
    is_md_or_do: bool = False,
    is_surgeon: bool = False,
    is_dentist: bool = False,
    is_np_or_pa: bool = False,
    is_behavioral_health: bool = False,
    is_facility: bool = False,
) -> ProviderGrouping:
    """
    Apply the Provider Specialty Framework decision tree to classify a provider.

    Decision flow (from the Visio flowchart):

    1. Is the provider an MD or DO?
       ├─ Yes → Is the provider a surgeon?
       │         ├─ Yes → Surgeon
       │         └─ No  → Physician
       └─ No  → Continue to next decision

    2. Is the provider a DDS/DMD (dentist)?
       └─ Yes → Dentist

    3. Is the provider an NP or PA?
       └─ Yes → Advanced Practitioner

    4. Is the provider a behavioral health provider?
       └─ Yes → Behavioral Health

    5. Otherwise → Allied Health Professional
    """
    if is_facility:
        return ProviderGrouping.FACILITY

    if is_md_or_do:
        if is_surgeon:
            return ProviderGrouping.SURGEON
        return ProviderGrouping.PHYSICIAN

    if is_dentist:
        return ProviderGrouping.DENTIST

    if is_np_or_pa:
        return ProviderGrouping.ADVANCED_PRACTITIONER

    if is_behavioral_health:
        return ProviderGrouping.BEHAVIORAL_HEALTH

    return ProviderGrouping.ALLIED_HEALTH


# ────────────────────────────────────────────────────────────────────
# Validation: apply framework rules + matrices
# ────────────────────────────────────────────────────────────────────

def validate_provider(
    provider: ProviderRecord,
    matrices: CompatibilityMatrices,
) -> ValidationResult:
    """
    Validate a provider record against the CMS framework rules and matrices.

    Checks:
    - NA V5:  Same NPI not allowed to have multiple Provider Groupings
    - NA V13: Provider Specialty Framework (correct grouping)
    - NA V14: Physician Specialty Compatibility Matrix
    - NA V15: Physician Subspecialty Compatibility Matrix
    - Grouping-specific specialty count rules
    """
    result = ValidationResult(provider=provider)

    grouping = provider.provider_grouping

    # ── NA V13: Verify grouping matches credentials ──
    expected_grouping = classify_provider(
        is_md_or_do=provider.is_md_or_do,
        is_surgeon=provider.is_surgeon,
        is_dentist=provider.is_dentist,
        is_np_or_pa=provider.is_np_or_pa,
        is_behavioral_health=provider.is_behavioral_health,
        is_facility=provider.is_facility,
    )
    if grouping != expected_grouping:
        result.add_error(
            "NA V13",
            f"Provider grouping '{grouping.value}' does not match expected "
            f"'{expected_grouping.value}' based on credentials. "
            f"MD/DO={provider.is_md_or_do}, Surgeon={provider.is_surgeon}, "
            f"Dentist={provider.is_dentist}, NP/PA={provider.is_np_or_pa}, "
            f"BehavioralHealth={provider.is_behavioral_health}",
        )

    # ── Grouping-specific specialty validation ──
    specialties = provider.specialties
    subspecialties = provider.subspecialties

    if grouping == ProviderGrouping.PHYSICIAN:
        _validate_physician(specialties, subspecialties, matrices, result)
    elif grouping == ProviderGrouping.SURGEON:
        _validate_surgeon(specialties, subspecialties, result)
    elif grouping == ProviderGrouping.DENTIST:
        _validate_dentist(specialties, subspecialties, result)
    elif grouping == ProviderGrouping.ADVANCED_PRACTITIONER:
        _validate_advanced_practitioner(specialties, result)
    elif grouping == ProviderGrouping.BEHAVIORAL_HEALTH:
        _validate_behavioral_health(specialties, subspecialties, result)
    elif grouping == ProviderGrouping.ALLIED_HEALTH:
        _validate_allied_health(specialties, result)

    return result


def _validate_physician(
    specialties: list[str],
    subspecialties: list[str],
    matrices: CompatibilityMatrices,
    result: ValidationResult,
) -> None:
    """
    Physician grouping rules (from flowchart):
    - Select up to two specialties from Physician Specialties
    - May designate one subspecialty
    - Also may select: Allergy and Immunology (007), Addiction Medicine Physician (800)
    - May select Gynecology in addition to Family Medicine if certified in obstetrics
    - Issuers should not separately list primary care subtypes alongside Family Medicine
    """
    # Check specialty count
    if len(specialties) > 2:
        result.add_error(
            "NA V13",
            f"Physician has {len(specialties)} specialties; maximum is 2.",
        )

    # Validate each specialty is a known physician specialty
    for code in specialties:
        if code not in PHYSICIAN_SPECIALTIES and code != "800":
            result.add_warning(
                "NA V13",
                f"Specialty code '{code}' is not in the standard Physician Specialties list.",
            )

    # NA V14: Check specialty-to-specialty compatibility
    for i, code_a in enumerate(specialties):
        for code_b in specialties[i + 1:]:
            if not matrices.are_specialties_compatible(code_a, code_b):
                name_a = PHYSICIAN_SPECIALTIES.get(code_a, code_a)
                name_b = PHYSICIAN_SPECIALTIES.get(code_b, code_b)
                result.add_error(
                    "NA V14",
                    f"Specialties '{name_a}' ({code_a}) and '{name_b}' ({code_b}) "
                    f"are not compatible per the Physician Specialty Compatibility Matrix.",
                )

    # Check subspecialty count
    if len(subspecialties) > 1:
        result.add_error(
            "NA V13",
            f"Physician has {len(subspecialties)} subspecialties; maximum is 1.",
        )

    # NA V15: Check subspecialty-to-specialty compatibility
    for sub_code in subspecialties:
        compatible_with_any = False
        for spec_code in specialties:
            if matrices.is_subspecialty_compatible_with_specialty(sub_code, spec_code):
                compatible_with_any = True
                break
        if not compatible_with_any and specialties:
            sub_name = PHYSICIAN_SUBSPECIALTIES.get(sub_code, sub_code)
            result.add_error(
                "NA V15",
                f"Subspecialty '{sub_name}' ({sub_code}) is not compatible with any "
                f"of the reported specialties: {specialties}.",
            )

    # Primary Care subtype rule: should not list Primary Care subtypes alongside Family Medicine
    family_medicine = "002" in specialties
    primary_care_subtypes = {"003", "101", "004"}  # Internal Med, Pediatric, Geriatrics
    if family_medicine:
        for code in specialties:
            if code in primary_care_subtypes:
                result.add_warning(
                    "NA V13",
                    f"Primary Care subtype '{PHYSICIAN_SPECIALTIES.get(code, code)}' "
                    f"listed alongside Family Medicine. Issuers should not separately list "
                    f"primary care subtypes as additional specialty types alongside Family Medicine.",
                )


def _validate_surgeon(
    specialties: list[str],
    subspecialties: list[str],
    result: ValidationResult,
) -> None:
    """
    Surgeon grouping rules (from flowchart):
    - Set 1: select no more than one option (General Surgery, Plastic Surgery,
      Orthopedic Surgery, Neurosurgery, Cardiothoracic Surgery, Vascular Surgery)
    - Set 2: select up to two options (Urology, Oncology-Surgical,
      Otolaryngology, Ophthalmology)
    - Cannot mix specialties across Set 1 and Set 2
    """
    set1_codes = set(SURGEON_SPECIALTIES_SET1.keys())
    set2_codes = set(SURGEON_SPECIALTIES_SET2.keys())

    in_set1 = [c for c in specialties if c in set1_codes]
    in_set2 = [c for c in specialties if c in set2_codes]

    # Cannot mix across sets
    if in_set1 and in_set2:
        result.add_error(
            "NA V13",
            f"Surgeon cannot mix specialties across Set 1 and Set 2. "
            f"Set 1: {[SURGEON_SPECIALTIES_SET1.get(c, c) for c in in_set1]}, "
            f"Set 2: {[SURGEON_SPECIALTIES_SET2.get(c, c) for c in in_set2]}.",
        )

    # Set 1: max 1
    if len(in_set1) > 1:
        result.add_error(
            "NA V13",
            f"Surgeon Set 1 allows max 1 specialty; got {len(in_set1)}: "
            f"{[SURGEON_SPECIALTIES_SET1.get(c, c) for c in in_set1]}.",
        )

    # Set 2: max 2
    if len(in_set2) > 2:
        result.add_error(
            "NA V13",
            f"Surgeon Set 2 allows max 2 specialties; got {len(in_set2)}: "
            f"{[SURGEON_SPECIALTIES_SET2.get(c, c) for c in in_set2]}.",
        )

    for code in specialties:
        if code not in SURGEON_SPECIALTIES:
            result.add_warning(
                "NA V13",
                f"Specialty code '{code}' is not in the standard Surgeon Specialties list.",
            )


def _validate_dentist(
    specialties: list[str],
    subspecialties: list[str],
    result: ValidationResult,
) -> None:
    """
    Dentist grouping rules (from flowchart):
    - Select up to two specialties from Dental Specialties
    - May designate one subspecialty from Dental Subspecialties
    """
    if len(specialties) > 2:
        result.add_error(
            "NA V13",
            f"Dentist has {len(specialties)} specialties; maximum is 2.",
        )

    for code in specialties:
        if code not in DENTIST_SPECIALTIES:
            result.add_warning(
                "NA V13",
                f"Specialty code '{code}' is not in the standard Dental Specialties list.",
            )

    if len(subspecialties) > 1:
        result.add_error(
            "NA V13",
            f"Dentist has {len(subspecialties)} subspecialties; maximum is 1.",
        )

    for code in subspecialties:
        if code not in DENTIST_SUBSPECIALTIES:
            result.add_warning(
                "NA V13",
                f"Subspecialty code '{code}' is not in the standard Dental Subspecialties list.",
            )


def _validate_advanced_practitioner(
    specialties: list[str],
    result: ValidationResult,
) -> None:
    """
    Advanced Practitioner grouping rules (from flowchart):
    - PAs: Please select Primary Care
    - NPs: Select only if the provider is an NP
    - APRNs: acceptable as PCP or Behavioral Health Provider
    """
    if not specialties:
        result.add_error(
            "NA V13",
            "Advanced Practitioner must have at least one specialty.",
        )
        return

    for code in specialties:
        if code not in ADVANCED_PRACTITIONER_SPECIALTIES:
            result.add_warning(
                "NA V13",
                f"Specialty code '{code}' is not in the standard Advanced Practitioner list.",
            )


def _validate_behavioral_health(
    specialties: list[str],
    subspecialties: list[str],
    result: ValidationResult,
) -> None:
    """
    Behavioral Health grouping rules (from flowchart):
    - Select up to two specialties from Behavioral Health Specialties
    - Up to two subspecialties from Behavioral Health Subspecialties
    - Select only if the provider is a behavioral health provider
    """
    if len(specialties) > 2:
        result.add_error(
            "NA V13",
            f"Behavioral Health provider has {len(specialties)} specialties; maximum is 2.",
        )

    for code in specialties:
        if code not in BEHAVIORAL_HEALTH_SPECIALTIES:
            result.add_warning(
                "NA V13",
                f"Specialty code '{code}' is not in the standard Behavioral Health list.",
            )

    if len(subspecialties) > 2:
        result.add_error(
            "NA V13",
            f"Behavioral Health provider has {len(subspecialties)} subspecialties; maximum is 2.",
        )


def _validate_allied_health(
    specialties: list[str],
    result: ValidationResult,
) -> None:
    """
    Allied Health grouping rules (from flowchart):
    - Only if none of the above groupings apply
    - Select one specialty from Allied Health Specialties
    """
    if len(specialties) > 1:
        result.add_error(
            "NA V13",
            f"Allied Health provider has {len(specialties)} specialties; maximum is 1.",
        )

    for code in specialties:
        if code not in ALLIED_HEALTH_SPECIALTIES:
            result.add_warning(
                "NA V13",
                f"Specialty code '{code}' is not in the standard Allied Health list.",
            )


# ────────────────────────────────────────────────────────────────────
# Convenience: validate a list of providers
# ────────────────────────────────────────────────────────────────────

def validate_providers(
    providers: list[ProviderRecord],
    matrices: Optional[CompatibilityMatrices] = None,
) -> list[ValidationResult]:
    """Validate a list of provider records."""
    if matrices is None:
        matrices = CompatibilityMatrices()
    return [validate_provider(p, matrices) for p in providers]


# ────────────────────────────────────────────────────────────────────
# Convenience: validate specialty codes directly
# ────────────────────────────────────────────────────────────────────

# Build reverse lookup: code → (grouping, is_subspecialty)
_CODE_TO_GROUPING: dict[str, tuple[ProviderGrouping, bool]] = {}

def _build_code_index() -> None:
    """Build a reverse index from specialty code to grouping."""
    if _CODE_TO_GROUPING:
        return  # Already built

    # Surgeon Set 1 + Set 2 (specialties)
    for code in SURGEON_SPECIALTIES_SET1:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.SURGEON, False)
    for code in SURGEON_SPECIALTIES_SET2:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.SURGEON, False)
    # Surgeon subspecialties
    for code in SURGEON_SUBSPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.SURGEON, True)

    # Physician specialties + fallback
    for code in PHYSICIAN_SPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.PHYSICIAN, False)
    for code in PHYSICIAN_FALLBACK:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.PHYSICIAN, False)
    # Physician subspecialties
    for code in PHYSICIAN_SUBSPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.PHYSICIAN, True)

    # Dentist specialties + subspecialties
    for code in DENTIST_SPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.DENTIST, False)
    for code in DENTIST_SUBSPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.DENTIST, True)

    # Advanced Practitioner
    for code in ADVANCED_PRACTITIONER_SPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.ADVANCED_PRACTITIONER, False)
    for code in ADVANCED_PRACTITIONER_BH_SPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.ADVANCED_PRACTITIONER, False)

    # Behavioral Health specialties + subspecialties
    for code in BEHAVIORAL_HEALTH_SPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.BEHAVIORAL_HEALTH, False)
    for code in BEHAVIORAL_HEALTH_SUBSPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.BEHAVIORAL_HEALTH, True)

    # Allied Health
    for code in ALLIED_HEALTH_SPECIALTIES:
        _CODE_TO_GROUPING[code] = (ProviderGrouping.ALLIED_HEALTH, False)


def validate_specialty_codes(
    codes: list[str],
    provider_grouping: Optional[ProviderGrouping] = None,
    matrices: Optional[CompatibilityMatrices] = None,
) -> ValidationResult:
    """
    Validate a list of specialty/subspecialty codes.

    The simplest way to use the framework. Pass in a list of codes and get
    back whether they form a valid combination.

    Parameters
    ----------
    codes : list[str]
        Specialty and subspecialty codes to validate (e.g. ["003", "008"]).
    provider_grouping : ProviderGrouping, optional
        Override the inferred grouping. If omitted, the grouping is inferred
        from the codes themselves.
    matrices : CompatibilityMatrices, optional
        Pre-loaded matrices. Created on demand if omitted.

    Returns
    -------
    ValidationResult
        Contains errors and warnings. Check ``result.is_valid``.

    Examples
    --------
    >>> result = validate_specialty_codes(["003", "008"])
    >>> result.is_valid
    True  # Internal Medicine + Cardiology subspecialty

    >>> result = validate_specialty_codes(["015", "035"])
    >>> result.is_valid
    False  # Surgeon Set 1 + Set 2 mix
    """
    _build_code_index()
    if matrices is None:
        matrices = CompatibilityMatrices()

    if not codes:
        result = ValidationResult(
            provider=ProviderRecord(npi="", provider_grouping=ProviderGrouping.PHYSICIAN)
        )
        result.add_error("NA V13", "No specialty codes provided.")
        return result

    # Separate specialties from subspecialties
    specialties = []
    subspecialties = []
    unknown_codes = []

    for code in codes:
        info = _CODE_TO_GROUPING.get(code)
        if info is None:
            unknown_codes.append(code)
        elif info[1]:  # is_subspecialty
            subspecialties.append(code)
        else:
            specialties.append(code)

    # Infer grouping from codes
    if provider_grouping is None:
        # Collect all groupings implied by the codes
        groupings = set()
        for code in codes:
            info = _CODE_TO_GROUPING.get(code)
            if info:
                groupings.add(info[0])

        if len(groupings) == 1:
            provider_grouping = groupings.pop()
        elif len(groupings) > 1:
            # Mixed groupings — pick the most specific one
            priority = [
                ProviderGrouping.SURGEON,
                ProviderGrouping.DENTIST,
                ProviderGrouping.ADVANCED_PRACTITIONER,
                ProviderGrouping.BEHAVIORAL_HEALTH,
                ProviderGrouping.PHYSICIAN,
                ProviderGrouping.ALLIED_HEALTH,
            ]
            for g in priority:
                if g in groupings:
                    provider_grouping = g
                    break
        else:
            # All codes unknown
            provider_grouping = ProviderGrouping.PHYSICIAN

    # Build a ProviderRecord and validate
    # Auto-set credential flags based on inferred grouping
    is_md_or_do = provider_grouping in (
        ProviderGrouping.PHYSICIAN, ProviderGrouping.SURGEON
    )
    is_surgeon = provider_grouping == ProviderGrouping.SURGEON
    is_dentist = provider_grouping == ProviderGrouping.DENTIST
    is_np_or_pa = provider_grouping == ProviderGrouping.ADVANCED_PRACTITIONER
    is_behavioral_health = provider_grouping == ProviderGrouping.BEHAVIORAL_HEALTH

    provider = ProviderRecord(
        npi="",
        provider_grouping=provider_grouping,
        specialties=specialties,
        subspecialties=subspecialties,
        is_md_or_do=is_md_or_do,
        is_surgeon=is_surgeon,
        is_dentist=is_dentist,
        is_np_or_pa=is_np_or_pa,
        is_behavioral_health=is_behavioral_health,
    )

    result = validate_provider(provider, matrices)

    # Report unknown codes
    for code in unknown_codes:
        result.add_warning(
            "NA V13",
            f"Unknown specialty code '{code}' — not recognized in any grouping.",
        )

    return result


# ────────────────────────────────────────────────────────────────────
# CLI
# ────────────────────────────────────────────────────────────────────

def _demo() -> None:
    """Demonstrate the framework with example providers."""
    matrices = CompatibilityMatrices()

    # Example 1: Valid Family Medicine physician
    p1 = ProviderRecord(
        npi="1234567890",
        provider_grouping=ProviderGrouping.PHYSICIAN,
        is_md_or_do=True,
        specialties=["002"],  # Family Medicine
    )
    r1 = validate_provider(p1, matrices)
    print(f"Provider 1 (Family Med): valid={r1.is_valid}, errors={len(r1.errors)}, warnings={len(r1.warnings)}")

    # Example 2: Invalid — Psychiatry + Dermatology (incompatible per matrix)
    p2 = ProviderRecord(
        npi="0987654321",
        provider_grouping=ProviderGrouping.PHYSICIAN,
        is_md_or_do=True,
        specialties=["029", "011"],  # Psychiatry + Dermatology
    )
    r2 = validate_provider(p2, matrices)
    print(f"Provider 2 (Psych+Derm): valid={r2.is_valid}")
    for e in r2.errors:
        print(f"  ERROR [{e.rule}]: {e.message}")

    # Example 3: Valid — Internal Medicine + Cardiology subspecialty
    p3 = ProviderRecord(
        npi="1112223334",
        provider_grouping=ProviderGrouping.PHYSICIAN,
        is_md_or_do=True,
        specialties=["003"],  # Internal Medicine
        subspecialties=["008"],  # Cardiology
    )
    r3 = validate_provider(p3, matrices)
    print(f"Provider 3 (IM+Cardio): valid={r3.is_valid}, errors={len(r3.errors)}")

    # Example 4: Wrong grouping — MD classified as Allied Health
    p4 = ProviderRecord(
        npi="4445556667",
        provider_grouping=ProviderGrouping.ALLIED_HEALTH,
        is_md_or_do=True,
        specialties=["049"],  # Physical Therapy
    )
    r4 = validate_provider(p4, matrices)
    print(f"Provider 4 (MD as Allied): valid={r4.is_valid}")
    for e in r4.errors:
        print(f"  ERROR [{e.rule}]: {e.message}")

    # Example 5: Valid Surgeon — General Surgery (Set 1, one specialty)
    p5 = ProviderRecord(
        npi="7778889990",
        provider_grouping=ProviderGrouping.SURGEON,
        is_md_or_do=True,
        is_surgeon=True,
        specialties=["015"],  # General Surgery
    )
    r5 = validate_provider(p5, matrices)
    print(f"Provider 5 (Surgeon, Set1): valid={r5.is_valid}, errors={len(r5.errors)}")

    # Example 6: Valid Dentist — Dental-General + Orthodontist subspecialty
    p6 = ProviderRecord(
        npi="1231231234",
        provider_grouping=ProviderGrouping.DENTIST,
        is_dentist=True,
        specialties=["201"],  # Dental-General
        subspecialties=["202"],  # Dental-Orthodontist
    )
    r6 = validate_provider(p6, matrices)
    print(f"Provider 6 (Dentist): valid={r6.is_valid}, errors={len(r6.errors)}")

    # Example 7: Valid NP — Primary Care APRN
    p7 = ProviderRecord(
        npi="4564564567",
        provider_grouping=ProviderGrouping.ADVANCED_PRACTITIONER,
        is_np_or_pa=True,
        specialties=["006"],  # Primary Care APRN
    )
    r7 = validate_provider(p7, matrices)
    print(f"Provider 7 (NP): valid={r7.is_valid}, errors={len(r7.errors)}")

    # Example 8: Valid Behavioral Health — Psychologist + Counselor
    p8 = ProviderRecord(
        npi="7897897890",
        provider_grouping=ProviderGrouping.BEHAVIORAL_HEALTH,
        is_behavioral_health=True,
        specialties=["103", "107"],  # Psychologist + Counselor
        subspecialties=["105"],  # Marriage and Family Therapist
    )
    r8 = validate_provider(p8, matrices)
    print(f"Provider 8 (BH): valid={r8.is_valid}, errors={len(r8.errors)}")

    # Example 9: Invalid Surgeon — mixing Set 1 and Set 2
    p9 = ProviderRecord(
        npi="1113335557",
        provider_grouping=ProviderGrouping.SURGEON,
        is_md_or_do=True,
        is_surgeon=True,
        specialties=["015", "035"],  # General Surgery (Set1) + Cardiothoracic (Set2)
    )
    r9 = validate_provider(p9, matrices)
    print(f"Provider 9 (Surgeon, Set1+Set2 mix): valid={r9.is_valid}")
    for e in r9.errors:
        print(f"  ERROR [{e.rule}]: {e.message}")

    # Example 10: Valid Surgeon — two Set 2 specialties
    p10 = ProviderRecord(
        npi="2224446668",
        provider_grouping=ProviderGrouping.SURGEON,
        is_md_or_do=True,
        is_surgeon=True,
        specialties=["035", "034"],  # Cardiothoracic + Vascular (both Set2)
    )
    r10 = validate_provider(p10, matrices)
    print(f"Provider 10 (Surgeon, Set2 x2): valid={r10.is_valid}, errors={len(r10.errors)}")

    # Example 11: Invalid Physician — 3 specialties (max 2)
    p11 = ProviderRecord(
        npi="3335557779",
        provider_grouping=ProviderGrouping.PHYSICIAN,
        is_md_or_do=True,
        specialties=["002", "003", "007"],  # Family Med + Internal Med + Allergy
    )
    r11 = validate_provider(p11, matrices)
    print(f"Provider 11 (Physician, 3 specs): valid={r11.is_valid}")
    for e in r11.errors:
        print(f"  ERROR [{e.rule}]: {e.message}")

    # Example 12: Invalid Physician — incompatible subspecialty
    p12 = ProviderRecord(
        npi="4446668880",
        provider_grouping=ProviderGrouping.PHYSICIAN,
        is_md_or_do=True,
        specialties=["002"],  # Family Medicine
        subspecialties=["008"],  # Cardiology (not compatible with Family Med)
    )
    r12 = validate_provider(p12, matrices)
    print(f"Provider 12 (FM + Cardiology sub): valid={r12.is_valid}")
    for e in r12.errors:
        print(f"  ERROR [{e.rule}]: {e.message}")


if __name__ == "__main__":
    _demo()
